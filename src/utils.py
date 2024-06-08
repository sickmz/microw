import json
import os

import gspread
from config import TELEGRAM_BOT_TOKEN, TELEGRAM_USER_ID
from constants import (
    LOCAL_BUDGET_PATH,
    LOCAL_CHART_PATH,
    LOCAL_EXPENSE_PATH,
    LOCAL_SETTINGS_PATH,
    categories,
)
from openpyxl import Workbook, load_workbook
from telegram import Bot, KeyboardButton, ReplyKeyboardMarkup

bot = Bot(token=TELEGRAM_BOT_TOKEN)


def build_keyboard(options, buttons_per_row=3):
    """
    Create a dynamic keyboard with a given number of buttons per row.
    """
    keyboard = []
    row = []
    for i, option in enumerate(options):
        row.append(KeyboardButton(option))
        if (i + 1) % buttons_per_row == 0 or i == len(options) - 1:
            keyboard.append(row)
            row = []
    return ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)


def load_settings():
    """
    Load settings from a JSON file. If the file doesn't exist, create it with default settings.
    """
    default_settings = {
        "google_sync": {"enabled": False, "last_upload": None},
        "budget_notifications": {"enabled": False},
    }
    if not os.path.exists(LOCAL_SETTINGS_PATH):
        settings = default_settings
        with open(LOCAL_SETTINGS_PATH, "w") as f:
            json.dump(settings, f)
    else:
        with open(LOCAL_SETTINGS_PATH, "r") as f:
            settings = json.load(f)
            if "google_sync" not in settings:
                settings["google_sync"] = default_settings["google_sync"]
            if "budget_notifications" not in settings:
                settings["budget_notifications"] = default_settings[
                    "budget_notifications"
                ]
            save_settings(settings)
    return settings


def save_settings(settings):
    """
    Save the given settings to a JSON file.
    """
    with open(LOCAL_SETTINGS_PATH, "w") as f:
        json.dump(settings, f)


def ensure_expense_file():
    """
    Ensure the expense file exists with appropriate headers.
    """
    if not os.path.exists(LOCAL_EXPENSE_PATH):
        os.makedirs((os.path.dirname(LOCAL_EXPENSE_PATH)), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.append(["Month", "Category", "Subcategory", "Price", "Date", "Timestamp"])
        wb.save(LOCAL_EXPENSE_PATH)


def ensure_budget_file():
    """
    Ensure the budget file exists with appropriate headers and initialize categories to 0.
    """
    if not os.path.exists(LOCAL_BUDGET_PATH):
        os.makedirs(os.path.dirname((LOCAL_BUDGET_PATH)), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.append(["Category", "Budget", "Spent"])
        for category in categories.keys():
            ws.append([category, 0, 0])
        wb.save(LOCAL_BUDGET_PATH)


def get_local_expense_wb():
    """
    Load the workbook and active sheet, ensuring the file and path exist.
    """
    ensure_expense_file()
    wb = load_workbook(LOCAL_EXPENSE_PATH)
    ws = wb.active
    return wb, ws


def get_local_budget_wb():
    """
    Load the budget workbook and active sheet, ensuring the file exists.
    """
    ensure_budget_file()
    wb = load_workbook(LOCAL_BUDGET_PATH)
    ws = wb.active
    return wb, ws


def get_remote_expense_wb(remote_spreadsheet_id, remote_worksheet_name):
    """
    Get the specified worksheet from a Google Sheets spreadsheet using its ID and worksheet name.
    Requires credentials stored in 'credentials.json'.
    """
    gc = gspread.service_account(filename="credentials.json")
    return gc.open_by_key(remote_spreadsheet_id).worksheet(remote_worksheet_name)


def ensure_charts_path():
    """
    Ensure the directory for storing charts exists, creating it if necessary.
    """
    if not os.path.exists(LOCAL_CHART_PATH):
        os.makedirs((LOCAL_CHART_PATH), exist_ok=True)


def is_local_expense_file_empty():
    """
    Check if the local expense file is empty or contains only the header row.
    """
    wb, ws = get_local_expense_wb()
    return ws.max_row <= 1


def set_budget(category, budget):
    """
    Set the budget for a given category.
    """
    wb, ws = get_local_budget_wb()
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=False):
        if row[0].value == category:
            row[1].value = budget
            wb.save(LOCAL_BUDGET_PATH)
            return
    ws.append([category, budget, 0])
    wb.save(LOCAL_BUDGET_PATH)


def get_budget(category):
    """
    Get the budget and spent amount for a given category.
    """
    ensure_budget_file()
    wb, ws = get_local_budget_wb()
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        if row[0] == category:
            return row[1], row[2]
    return 0, 0


def update_spent(category, amount):
    """
    Update the spent amount for a given category.
    """
    ensure_budget_file()
    wb, ws = get_local_budget_wb()
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=False):
        if row[0].value == category:
            row[2].value += amount
            wb.save(LOCAL_BUDGET_PATH)
            return
    ws.append([category, 0, amount])
    wb.save(LOCAL_BUDGET_PATH)


async def check_budget(category):
    """
    Notify the user if the spending for the given category exceeds the budget.
    """
    settings = load_settings()
    if not settings["budget_notifications"]["enabled"]:
        return

    budget, spent = get_budget(category)
    if spent > budget & budget > 0:
        message = (
            f"Alert ⚠️ \n\nBudget exceeded for <u>{category}</u>\n"
            f"You spent {spent} € and your budget was {budget} € \n"
            f"You exceeded your budget by <b>{spent - budget}</b> €"
        )
        await bot.send_message(
            chat_id=TELEGRAM_USER_ID, text=message, parse_mode="HTML"
        )


def get_current_budget(category: str) -> float:
    wb, ws = get_local_budget_wb()
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        if row[0] == category:
            return row[1]
    return 0.0
