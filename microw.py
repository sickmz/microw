import logging
import os
import json
import pandas as pd
import matplotlib.pyplot as plt
import datetime
import calendar
import time
from apscheduler.schedulers.background import BackgroundScheduler
import gspread
import seaborn as sns
from cachetools import TTLCache
from telegram import Update, ReplyKeyboardMarkup, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, ContextTypes, ConversationHandler, MessageHandler, CallbackQueryHandler, filters
from dotenv import load_dotenv, dotenv_values
from openpyxl import load_workbook, Workbook

# Configure logging
logging.basicConfig(format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()
env_vars = dotenv_values()
BOT_TOKEN = env_vars.get("BOT_TOKEN")
USER_ID = env_vars.get("USER_ID")
SPREADSHEET_ID = env_vars.get("SPREADSHEET_ID")
EXPENSE_SHEET = env_vars.get("EXPENSE_SHEET")
LOCAL_EXPENSE_PATH = env_vars.get("LOCAL_EXPENSE_PATH")

# Set up cache
cache = TTLCache(maxsize=100, ttl=86400)

# Constants for conversation steps
CHOOSING, CHOOSING_CATEGORY, CHOOSING_SUBCATEGORY, CHOOSING_PRICE, CHOOSING_ITEM_TO_DELETE, CHOOSING_CHART = range(6)

# Define reply keyboard
reply_keyboard = [
    ["✏️ Add", "❌ Delete", "📊 Charts"],
    ["📋 List", "🔄 Reset", "⚙️ Settings"],
]
markup = ReplyKeyboardMarkup(reply_keyboard, one_time_keyboard=False, resize_keyboard=True)

# Define categories
categories = {
    'Home': ['Gas', 'Light', 'Water', 'Tari', 'Rent', 'Products'],
    'Food': ['Market', 'Delivery', 'Gastronomy'],
    'Subscriptions': ['Fastweb', 'Ho', 'Everli', 'Prime'],
}

# Helper function to build keyboard
def build_keyboard(options: list) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[InlineKeyboardButton(option, callback_data=option)] for option in options])

# Ensure the spreadsheet directory and file exist
def ensure_spreadsheet_path(file_path):
    if not os.path.exists(os.path.dirname(file_path)):
        os.makedirs(os.path.dirname(file_path))
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Month", "Category", "Subcategory", "Price", "Date", "Timestamp"])
        wb.save(file_path)

# Load or initialize workbook/worksheet
def get_workbook_and_sheet(file_path):
    key = f"local:{file_path}"
    if key not in cache:
        ensure_spreadsheet_path(file_path)
        wb = load_workbook(file_path)
        ws = wb.active
        cache[key] = (wb, ws)
    return cache[key]

# Start command handler
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if str(update.effective_user.id) != str(USER_ID):
        await update.message.reply_text("You're not authorized. ⛔")
        return ConversationHandler.END
    await update.effective_message.reply_text("Hi! I'm microw. What can I do for you?", reply_markup=markup)
    return CHOOSING

# Ask for category
async def ask_category(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("Select a category:", reply_markup=build_keyboard(categories.keys()))
    return CHOOSING_CATEGORY

# Ask for subcategory
async def ask_subcategory(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['selected_category'] = update.callback_query.data
    await update.callback_query.message.edit_text("Select a subcategory:", reply_markup=build_keyboard(categories[context.user_data['selected_category']]))
    return CHOOSING_SUBCATEGORY

# Ask for price
async def ask_price(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["selected_subcategory"] = update.callback_query.data
    await update.callback_query.message.edit_text("Enter the price for this item:")
    return CHOOSING_PRICE

# Save expense in the local spreadsheet with a timestamp
async def save_on_local_spreadsheet(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        price = float(update.message.text.replace(',', '.'))
        category = context.user_data["selected_category"]
        subcategory = context.user_data["selected_subcategory"]
        
        wb, ws = get_workbook_and_sheet(LOCAL_EXPENSE_PATH)

        record_timestamp = datetime.datetime.now().isoformat()
        ws.append([datetime.datetime.now().strftime("%B"), category, subcategory, price, datetime.datetime.now().strftime('%d/%m/%Y'), record_timestamp])
        wb.save(LOCAL_EXPENSE_PATH)
        await update.message.reply_text(f"<b>Expense saved 📌</b>\n\n<b>Category:</b> {category}\n<b>Subcategory:</b> {subcategory}\n<b>Price:</b> {price} €", parse_mode='HTML')
    except ValueError:
        await update.message.reply_text("Please enter a valid price. 🚨")
    return await start(update, context)

# Pagination
ITEMS_PER_PAGE = 5

# Starting process for deleting an expense
async def ask_deleting(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if 'current_page' not in context.user_data:
        context.user_data['current_page'] = 0
    return await show_expenses(update, context)

# Display expenses with navigation button
async def show_expenses(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    wb, ws = get_workbook_and_sheet(LOCAL_EXPENSE_PATH)
    expenses = pd.DataFrame(ws.values)
    
    if len(expenses.columns) > 0:
        expenses.columns = expenses.iloc[0]  
        expenses = expenses[1:] 
    
    num_rows = len(expenses)
    current_page = context.user_data['current_page']
    
    start_index = max(num_rows - (current_page + 1) * ITEMS_PER_PAGE, 0)
    end_index = min(num_rows - current_page * ITEMS_PER_PAGE, num_rows)

    expense_buttons = []
    for i, (index, row) in enumerate(expenses.iloc[start_index:end_index].iterrows()):
        expense_buttons.append([InlineKeyboardButton(
            f"🗑️ ({row['Date']}) {row['Subcategory']}: {row['Price']} €",
            callback_data=f"delete_{index}")]) 

    navigation_buttons = []
    if start_index > 0:
        navigation_buttons.append(InlineKeyboardButton("⬅️ Previous", callback_data="previous"))
    if end_index < num_rows:
        navigation_buttons.append(InlineKeyboardButton("➡️ Next", callback_data="next"))
    
    all_buttons = expense_buttons + [navigation_buttons] if navigation_buttons else expense_buttons
    
    if hasattr(update, 'callback_query') and update.callback_query:
        await update.callback_query.message.edit_text("Choose an expense to delete:", reply_markup=InlineKeyboardMarkup(all_buttons))
        await update.callback_query.answer()
    else:
        await update.message.reply_text("Choose an expense to delete:", reply_markup=InlineKeyboardMarkup(all_buttons))

    return CHOOSING_ITEM_TO_DELETE

async def handle_navigation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query_data = update.callback_query.data
    
    if query_data.startswith("delete_"):
        expense_id = int(query_data.split("_")[1])
        return await delete_expense(update, context, expense_id)
    elif query_data == "previous":
        context.user_data['current_page'] += 1
    elif query_data == "next":
        context.user_data['current_page'] -= 1
    
    return await show_expenses(update, context)

# Deleting an expense
async def delete_expense(update: Update, context: ContextTypes.DEFAULT_TYPE, expense_id: int) -> int:
    wb, ws = get_workbook_and_sheet(LOCAL_EXPENSE_PATH)
    try:
        ws.delete_rows(expense_id + 1)  # Adjust row index for header
        wb.save(LOCAL_EXPENSE_PATH)
        context.user_data['current_page'] = 0
        await update.callback_query.message.edit_text("Expense deleted successfully. ✅")
    except Exception as e:
        await update.callback_query.reply_text(f"Error deleting expense: {e}", show_alert=True)
    
    return await start(update, context)

async def make_charts(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    chart_buttons = [
        [InlineKeyboardButton("Expense by category (yearly)", callback_data="chart_yearly")],
        [InlineKeyboardButton("Expense by category (monthly)", callback_data="chart_monthly")],
        [InlineKeyboardButton("Trend top 3 categories (monthly)", callback_data="chart_trend")],
        [InlineKeyboardButton("Heatmap expense intensity (monthly)", callback_data="chart_heatmap")]
    ]
    await update.message.reply_text("Select a chart to view:", reply_markup=InlineKeyboardMarkup(chart_buttons))

    return CHOOSING_CHART

def ensure_charts_path():
    if not os.path.exists('charts'):
        os.makedirs('charts')

def is_expense_file_empty():
    wb, ws = get_workbook_and_sheet(LOCAL_EXPENSE_PATH)
    return ws.max_row <= 1

async def show_yearly_chart(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if is_expense_file_empty():
        await update.callback_query.message.reply_text("You have not yet registered expenses.")
        return await start(update, context)

    wb, ws = get_workbook_and_sheet(LOCAL_EXPENSE_PATH)
    values = pd.DataFrame(ws.values)
    if len(values.columns) > 0:
        values.columns = values.iloc[0]
        values = values[1:]
    df = pd.DataFrame(values, columns=values.columns)
    df['Price'] = df['Price'].astype(float)
    df['Date'] = pd.to_datetime(df['Date'], format='%d/%m/%Y')
    
    await save_pie_chart(df, 'charts/expense_by_category_by_year.png')
    await update.callback_query.message.reply_photo(open('charts/expense_by_category_by_year.png', 'rb'), caption="Expense by category (yearly)")

    return await start(update, context)

async def show_trend_chart(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if is_expense_file_empty():
        await update.callback_query.message.reply_text("You have not yet registered expenses.")
        return await start(update, context)
    wb, ws = get_workbook_and_sheet(LOCAL_EXPENSE_PATH)
    values = pd.DataFrame(ws.values)
    if len(values.columns) > 0:
        values.columns = values.iloc[0]
        values = values[1:]
    df = pd.DataFrame(values, columns=values.columns)
    df['Price'] = df['Price'].astype(float)
    df['Date'] = pd.to_datetime(df['Date'], format='%d/%m/%Y')

    await save_trend_chart(df, 'charts/expense_trend_top_categories_by_month.png')
    await update.callback_query.message.reply_photo(open('charts/expense_trend_top_categories_by_month.png', 'rb'), caption="Trend top 3 categories (monthly)")

    return await start(update, context)

async def show_monthly_chart(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if is_expense_file_empty():
        await update.callback_query.message.reply_text("You have not yet registered expenses.")
        return await start(update, context)
    wb, ws = get_workbook_and_sheet(LOCAL_EXPENSE_PATH)
    values = pd.DataFrame(ws.values)
    if len(values.columns) > 0:
        values.columns = values.iloc[0]
        values = values[1:]
    df = pd.DataFrame(values, columns=values.columns)
    df['Price'] = df['Price'].astype(float)
    df['Date'] = pd.to_datetime(df['Date'], format='%d/%m/%Y')

    await save_stacked_bar_chart(df, 'charts/monthly_expenses_by_category.png')
    await update.callback_query.message.reply_photo(open('charts/monthly_expenses_by_category.png', 'rb'), caption="Expense by category (monthly)")

    return await start(update, context)

async def show_heatmap_chart(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if is_expense_file_empty():
        await update.callback_query.message.reply_text("You have not yet registered expenses.")
        return await start(update, context)
    wb, ws = get_workbook_and_sheet(LOCAL_EXPENSE_PATH)
    values = pd.DataFrame(ws.values)
    if len(values.columns) > 0:
        values.columns = values.iloc[0]
        values = values[1:]
    df = pd.DataFrame(values, columns=values.columns)
    df['Price'] = df['Price'].astype(float)
    df['Date'] = pd.to_datetime(df['Date'], format='%d/%m/%Y')

    await save_heatmap(df, 'charts/heatmap_expense_intensity.png')
    await update.callback_query.message.reply_photo(open('charts/heatmap_expense_intensity.png', 'rb'), caption="Heatmap of expense intensity (monthly)")

    return await start(update, context)

async def save_pie_chart(df, filename):
    ensure_charts_path()  
    expenses_by_category = df.groupby('Category')['Price'].sum().reset_index()
    plt.figure(figsize=(10, 6))
    pie = plt.pie(expenses_by_category['Price'], autopct=lambda p: f'{p:.1f}% ({p*sum(expenses_by_category["Price"])/100:.2f} €)' if p > 5 else '', startangle=90)
    plt.legend(pie[0], expenses_by_category['Category'], loc="best")
    plt.axis('equal')
    plt.savefig(filename)
    plt.close()

async def save_trend_chart(df, filename):
    ensure_charts_path()  
    df['Month'] = df['Date'].dt.month
    top_categories = df.groupby('Category')['Price'].sum().nlargest(3).index
    top_categories_data = df[df['Category'].isin(top_categories)]
    expenses_by_month_category = top_categories_data.groupby(['Month', 'Category'])['Price'].sum().unstack(fill_value=0)
    plt.figure(figsize=(10, 6))
    month_names = [calendar.month_name[i] for i in range(1, 13)]
    ax = expenses_by_month_category.plot(kind='line', marker='o', ax=plt.gca())
    ax.set_xticks(range(1, 13))
    ax.set_xticklabels(month_names, rotation=45, ha='right')
    ax.set_xlabel('') 
    plt.legend(title='Category', loc='upper right')
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(filename)
    plt.close()

async def save_stacked_bar_chart(df, filename):
    ensure_charts_path()  
    df['Month'] = df['Date'].dt.strftime('%B')
    monthly_expenses = df.groupby(['Month', 'Category'])['Price'].sum().unstack().fillna(0)
    months_order = list(calendar.month_name[1:])
    monthly_expenses = monthly_expenses.reindex(months_order)
    plt.figure(figsize=(12, 8))
    ax = monthly_expenses.plot(kind='bar', stacked=True, width=0.8, zorder=3)
    ax.set_xticklabels(months_order, rotation=45, ha='right')
    ax.set_xlabel('') 
    plt.legend(loc='upper right')
    plt.grid(True, zorder=0)
    plt.tight_layout()
    plt.savefig(filename)
    plt.close()

async def save_heatmap(df, filename):
    ensure_charts_path()  
    df['Month'] = df['Date'].dt.strftime('%B')
    heatmap_data = df.pivot_table(values='Price', index='Category', columns='Month', aggfunc='sum', fill_value=0)
    existing_months = [month for month in calendar.month_name[1:] if month in heatmap_data.columns]
    heatmap_data = heatmap_data[existing_months]
    plt.figure(figsize=(12, 8))
    sns.heatmap(heatmap_data, fmt=".2f", annot=True, cmap="YlGnBu")
    plt.tight_layout()
    plt.savefig(filename)
    plt.close()

# List of expenses
async def make_list(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    wb, ws = get_workbook_and_sheet(LOCAL_EXPENSE_PATH)
    values = pd.DataFrame(ws.values)
    if len(values.columns) > 0:
        values.columns = values.iloc[0]
        values = values[1:]
    df = pd.DataFrame(values, columns=values.columns)
    df['Price'] = df['Price'].astype(float)
    df['Date'] = pd.to_datetime(df['Date'], format='%d/%m/%Y')
    current_year = datetime.datetime.now().year
    df_current_year = df[df['Date'].dt.year == current_year]

    message = ""
    grouped = df_current_year.groupby([df_current_year['Date'].dt.month, 'Category'])['Price'].sum()
    total_per_month = df_current_year.groupby(df_current_year['Date'].dt.month)['Price'].sum()
    for month in range(1, datetime.datetime.now().month + 1):
        month_name = calendar.month_name[month]
        message += f"\n<b>{month_name}:</b>\n"
        if month in grouped.index.get_level_values(0):
            for category, amount in grouped[month].items():
                message += f"  - {category}: {amount:.2f} €\n"
        total = total_per_month.get(month, 0)
        message += f"  <b>Total:</b> {total:.2f} €\n"
    await update.message.reply_text(message, parse_mode='HTML')

    return await start(update, context)

# Invalid action handler
async def invalid_transition(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("Invalid action. One thing at a time..")

    return await start(update, context)

# Reset conversation
async def fallback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()

    return await start(update, context)

# Load settings with last upload timestamp
def load_settings():
    if not os.path.exists('settings.json'):
        settings = {'google_sync_enabled': False, 'last_upload': None}
        with open('settings.json', 'w') as f:
            json.dump(settings, f)
    else:
        with open('settings.json', 'r') as f:
            settings = json.load(f)
    return settings

def save_settings(settings):
    with open('settings.json', 'w') as f:
        json.dump(settings, f)

# Options to enable or disable Google sync
async def ask_settings(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    settings = load_settings()
    sync_status = "enabled" if settings.get('google_sync_enabled', False) else "disabled"
    message = f"Google Sheets synchronization is currently <b>{sync_status}</b>. Choose an action:"
    
    settings_keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("Enable Google Sheet sync", callback_data="enable_sync")],
        [InlineKeyboardButton("Disable Google Sheet sync", callback_data="disable_sync")]
    ])
    
    await update.message.reply_text(message, reply_markup=settings_keyboard, parse_mode='HTML')
    return CHOOSING

# User choice to enable or disable Google sync
async def handle_settings_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query_data = update.callback_query.data
    settings = load_settings()
    
    if query_data == "enable_sync":
        settings['google_sync_enabled'] = True
        message = "Google Sheets synchronization is now enabled."
    elif query_data == "disable_sync":
        settings['google_sync_enabled'] = False
        message = "Google Sheets synchronization is now disabled."
    
    save_settings(settings)
    
    await update.callback_query.answer()
    await update.callback_query.message.edit_text(message)
    
    return CHOOSING


# Get Google Sheet worksheet
def get_worksheet(spreadsheet_id, worksheet_name):
    key = f"{spreadsheet_id}:{worksheet_name}"
    if key not in cache:
        gc = gspread.service_account(filename='credentials.json')
        cache[key] = gc.open_by_key(spreadsheet_id).worksheet(worksheet_name)

    return cache[key]

# Get Google Sheet worksheet
def get_worksheet(spreadsheet_id, worksheet_name):
    key = f"{spreadsheet_id}:{worksheet_name}"
    if key not in cache:
        gc = gspread.service_account(filename='credentials.json')
        cache[key] = gc.open_by_key(spreadsheet_id).worksheet(worksheet_name)

    return cache[key]

def sync_to_google_sheets():
    logger.info("Sync function started")
    settings = load_settings()
    if settings.get('google_sync_enabled'):
        logger.info("Google sync is enabled")
        wb, ws = get_workbook_and_sheet(LOCAL_EXPENSE_PATH)
        df = pd.DataFrame(ws.values)
        if len(df.columns) > 0:
            df.columns = df.iloc[0]
            df = df[1:]
        
        last_upload = pd.to_datetime(settings.get('last_upload'))
        new_records = df[pd.to_datetime(df['Timestamp']) > last_upload] if last_upload else df
        
        if not new_records.empty:
            logger.info(f"New records to upload: {len(new_records)}")
            sheet = get_worksheet(SPREADSHEET_ID, EXPENSE_SHEET)
            
            records_to_upload = new_records[['Month', 'Category', 'Subcategory', 'Price', 'Date']].values.tolist()
            for record in records_to_upload:
                sheet.append_row(record)
                logger.info(f"Uploaded record: {record}")
            
            new_last_upload = pd.to_datetime(new_records['Timestamp']).max()
            settings['last_upload'] = new_last_upload.isoformat()
            save_settings(settings)
        else:
            logger.info("No new records to upload")
    else:
        logger.info("Google sync is disabled")

# APScheduler to manage background tasks
def start_scheduler():
    scheduler = BackgroundScheduler()
    scheduler.add_job(sync_to_google_sheets, 'interval', minutes=60)
    scheduler.start()
    return scheduler

def main() -> None:
    application = Application.builder().token(BOT_TOKEN).build()
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            CHOOSING: [
                MessageHandler(filters.Regex("^✏️ Add$"), ask_category),
                MessageHandler(filters.Regex("^❌ Delete$"), ask_deleting),
                MessageHandler(filters.Regex("^📊 Charts$"), make_charts),
                MessageHandler(filters.Regex("^📋 List$"), make_list),
                MessageHandler(filters.Regex("^⚙️ Settings$"), ask_settings),
                CallbackQueryHandler(handle_settings_choice, pattern="^(enable_sync|disable_sync)$"),
            ],
            CHOOSING_CATEGORY: [
                CallbackQueryHandler(ask_subcategory), 
                MessageHandler(filters.Regex("^(❌ Delete|📊 Charts|📋 List|⚙️ Settings)$"), invalid_transition)],
            CHOOSING_SUBCATEGORY: [
                CallbackQueryHandler(ask_price), 
                MessageHandler(filters.Regex("^(❌ Delete|📊 Charts|📋 List|⚙️ Settings)$"), invalid_transition)],
            CHOOSING_PRICE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, save_on_local_spreadsheet), 
                MessageHandler(filters.Regex("^(❌ Delete|📊 Charts|📋 List|⚙️ Settings)$"), invalid_transition)],
            CHOOSING_ITEM_TO_DELETE: [
                CallbackQueryHandler(handle_navigation),
                MessageHandler(filters.Regex("^(✏️ Add|📊 Charts|📋 List|⚙️ Settings)$"), invalid_transition)],
            CHOOSING_CHART: [
                CallbackQueryHandler(show_yearly_chart, pattern="^chart_yearly$"),
                CallbackQueryHandler(show_monthly_chart, pattern="^chart_monthly$"),
                CallbackQueryHandler(show_trend_chart, pattern="^chart_trend$"),
                CallbackQueryHandler(show_heatmap_chart, pattern="^chart_heatmap$"),
                MessageHandler(filters.Regex("^(✏️ Add|❌ Delete|📋 List|⚙️ Settings)$"), invalid_transition)],         
        },
        fallbacks=[MessageHandler(filters.Regex("^🔄 Reset$"), fallback)],
    )
    application.add_handler(conv_handler)
    application.run_polling(allowed_updates=Update.ALL_TYPES)

    scheduler = start_scheduler()

if __name__ == "__main__":
    scheduler = start_scheduler()
    main()