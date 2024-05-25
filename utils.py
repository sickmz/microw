import os
import json
from openpyxl import load_workbook, Workbook
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
import gspread

from constants import categories
from constants import BUDGET_PATH, EXPENSE_PATH, SETTINGS_PATH

def build_keyboard(options: list) -> InlineKeyboardMarkup:
    """
    Build an inline keyboard.
    Each option becomes a button with the same label and callback data.
    """
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(option, callback_data=option)]
        for option in options
    ])


def is_expense_file_empty():
    """
    Check if the local expense file is empty or contains only the header row.
    """
    wb, ws = get_workbook_and_sheet(EXPENSE_PATH)
    return ws.max_row <= 1


def ensure_spreadsheet_path(file_path):
    """
    Ensure the directory for the spreadsheet exists and create the file with a header if it doesn't exist.
    """
    if not os.path.exists(os.path.dirname(file_path)):
        os.makedirs(os.path.dirname(file_path))
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Month", "Category", "Subcategory",
                  "Price", "Date", "Timestamp"])
        wb.save(file_path)


def get_workbook_and_sheet(file_path):
    """
    Load the workbook and active sheet, ensuring the file and path exist.
    """
    ensure_spreadsheet_path(file_path)
    wb = load_workbook(file_path)
    ws = wb.active
    return wb, ws


def load_settings():
    """
    Load settings from a JSON file. If the file doesn't exist, create it with default settings.
    """
    if not os.path.exists(SETTINGS_PATH):
        settings = {'google_sync_enabled': False, 'last_upload': None}
        with open(SETTINGS_PATH, 'w') as f:
            json.dump(settings, f)
    else:
        with open(SETTINGS_PATH, 'r') as f:
            settings = json.load(f)
    return settings


def save_settings(settings):
    """
    Save the given settings to a JSON file.
    """
    with open(SETTINGS_PATH, 'w') as f:
        json.dump(settings, f)


def get_worksheet(spreadsheet_id, worksheet_name):
    """
    Get the specified worksheet from a Google Sheets spreadsheet using its ID and worksheet name.
    Requires credentials stored in 'credentials.json'.
    """
    gc = gspread.service_account(filename='credentials.json')
    return gc.open_by_key(spreadsheet_id).worksheet(worksheet_name)


def ensure_charts_path():
    """
    Ensure the directory for storing charts exists, creating it if necessary.
    """
    if not os.path.exists('charts'):
        os.makedirs('charts')


def ensure_budget_file():
    """
    Ensure the budget file exists with appropriate headers and initialize categories to 0.
    """
    if not os.path.exists(BUDGET_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["Category", "Budget", "Spent"])
        for category in categories.keys():
            ws.append([category, 0, 0])
        wb.save(BUDGET_PATH)


def get_budget_workbook_and_sheet():
    """
    Load the budget workbook and active sheet, ensuring the file exists.
    """
    ensure_budget_file()
    wb = load_workbook(BUDGET_PATH)
    ws = wb.active
    return wb, ws

def set_budget(category, budget):
    """
    Set the budget for a given category.
    """
    wb, ws = get_budget_workbook_and_sheet()
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=False):
        if row[0].value == category:
            row[1].value = budget
            wb.save(BUDGET_PATH)
            return
    ws.append([category, budget, 0])
    wb.save(BUDGET_PATH)

def get_budget(category):
    """
    Get the budget and spent amount for a given category.
    """
    wb, ws = get_budget_workbook_and_sheet()
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        if row[0] == category:
            return row[1], row[2]
    return 0, 0

def update_spent(category, amount):
    """
    Update the spent amount for a given category.
    """
    wb, ws = get_budget_workbook_and_sheet()
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=False):
        if row[0].value == category:
            row[2].value += amount
            wb.save(BUDGET_PATH)
            return
    ws.append([category, 0, amount])
    wb.save(BUDGET_PATH)